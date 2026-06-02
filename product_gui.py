import sys
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QTableWidget, QTableWidgetItem,
    QSpinBox, QComboBox, QMessageBox, QFileDialog, QFrame
)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont, QColor
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from product_database import ProductDatabase
from datetime import datetime


class ProductGUI(QMainWindow):
    """PyQt5를 사용한 제품 관리 GUI"""
    
    def __init__(self):
        super().__init__()
        self.db = ProductDatabase("products.db")
        self.selected_product_id = None
        self.init_ui()
    
    def init_ui(self):
        """UI 초기화"""
        self.setWindowTitle("제품 관리 시스템")
        self.setGeometry(100, 100, 800, 600)
        
        # 전체 스타일 설정
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f5f5f5;
            }
            QLabel {
                color: #333333;
                font-weight: bold;
            }
            QLineEdit {
                padding: 8px;
                border: 2px solid #e0e0e0;
                border-radius: 5px;
                background-color: white;
                selection-background-color: #2196F3;
            }
            QLineEdit:focus {
                border: 2px solid #2196F3;
                background-color: #f9f9f9;
            }
            QSpinBox, QComboBox {
                padding: 8px;
                border: 2px solid #e0e0e0;
                border-radius: 5px;
                background-color: white;
                selection-background-color: #2196F3;
            }
            QSpinBox:focus, QComboBox:focus {
                border: 2px solid #2196F3;
                background-color: #f9f9f9;
            }
        """)
        
        # 중앙 위젯
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # 전체 레이아웃
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(15)
        main_layout.setContentsMargins(15, 15, 15, 15)
        
        # 상단 입력 영역
        input_layout = self.create_input_area()
        main_layout.addLayout(input_layout)
        
        # 중단 버튼 영역
        button_layout = self.create_button_area()
        main_layout.addLayout(button_layout)
        
        # 하단 테이블 영역
        table_layout = self.create_table_area()
        main_layout.addLayout(table_layout)
        
        # 초기 데이터 로드
        self.load_products()
        
        self.show()
    
    def create_input_area(self):
        """입력 영역 생성"""
        layout = QHBoxLayout()
        
        # 제목 라벨
        title_label = QLabel("📝 제품 입력 정보")
        title_font = QFont()
        title_font.setBold(True)
        title_font.setPointSize(11)
        title_label.setFont(title_font)
        title_label.setStyleSheet("color: #1976D2; padding: 5px;")
        layout.addWidget(title_label)
        layout.addStretch()
        
        # 제품명
        layout.addWidget(QLabel("제품명:"))
        self.input_name = QLineEdit()
        self.input_name.setPlaceholderText("제품 이름 입력")
        layout.addWidget(self.input_name)
        
        # 가격
        layout.addWidget(QLabel("가격:"))
        self.input_price = QSpinBox()
        self.input_price.setMaximum(10000000)
        self.input_price.setSingleStep(1000)
        layout.addWidget(self.input_price)
        
        # 재고
        layout.addWidget(QLabel("재고:"))
        self.input_stock = QSpinBox()
        self.input_stock.setMaximum(1000)
        layout.addWidget(self.input_stock)
        
        # 카테고리
        layout.addWidget(QLabel("카테고리:"))
        self.input_category = QComboBox()
        self.input_category.addItems(["전자제품", "액세서리", "도서", "의류", "기타"])
        layout.addWidget(self.input_category)
        
        return layout
    
    def create_button_area(self):
        """버튼 영역 생성"""
        layout = QHBoxLayout()
        
        # 버튼 스타일
        button_style = """
            QPushButton {
                padding: 10px 20px;
                border: none;
                border-radius: 6px;
                font-weight: bold;
                font-size: 12px;
                color: white;
                background-color: #2196F3;
                transition: background-color 0.3s;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
            QPushButton:pressed {
                background-color: #1565C0;
            }
        """
        
        update_button_style = """
            QPushButton {
                padding: 10px 20px;
                border: none;
                border-radius: 6px;
                font-weight: bold;
                font-size: 12px;
                color: white;
                background-color: #FF9800;
                transition: background-color 0.3s;
            }
            QPushButton:hover {
                background-color: #F57C00;
            }
            QPushButton:pressed {
                background-color: #E65100;
            }
        """
        
        delete_button_style = """
            QPushButton {
                padding: 10px 20px;
                border: none;
                border-radius: 6px;
                font-weight: bold;
                font-size: 12px;
                color: white;
                background-color: #F44336;
                transition: background-color 0.3s;
            }
            QPushButton:hover {
                background-color: #D32F2F;
            }
            QPushButton:pressed {
                background-color: #C62828;
            }
        """
        
        excel_button_style = """
            QPushButton {
                padding: 10px 20px;
                border: none;
                border-radius: 6px;
                font-weight: bold;
                font-size: 12px;
                color: white;
                background-color: #4CAF50;
                transition: background-color 0.3s;
            }
            QPushButton:hover {
                background-color: #388E3C;
            }
            QPushButton:pressed {
                background-color: #2E7D32;
            }
        """
        
        refresh_button_style = """
            QPushButton {
                padding: 10px 20px;
                border: none;
                border-radius: 6px;
                font-weight: bold;
                font-size: 12px;
                color: white;
                background-color: #9C27B0;
                transition: background-color 0.3s;
            }
            QPushButton:hover {
                background-color: #7B1FA2;
            }
            QPushButton:pressed {
                background-color: #6A1B9A;
            }
        """
        
        # 추가 버튼
        btn_add = QPushButton("추가")
        btn_add.clicked.connect(self.add_product)
        btn_add.setMinimumWidth(80)
        btn_add.setStyleSheet(button_style)
        layout.addWidget(btn_add)
        
        # 수정 버튼
        btn_update = QPushButton("수정")
        btn_update.clicked.connect(self.update_product)
        btn_update.setMinimumWidth(80)
        btn_update.setStyleSheet(update_button_style)
        layout.addWidget(btn_update)
        
        # 삭제 버튼
        btn_delete = QPushButton("삭제")
        btn_delete.clicked.connect(self.delete_product)
        btn_delete.setMinimumWidth(80)
        btn_delete.setStyleSheet(delete_button_style)
        layout.addWidget(btn_delete)
        
        layout.addStretch()
        
        # 엑셀 저장 버튼
        btn_excel = QPushButton("엑셀 저장")
        btn_excel.clicked.connect(self.save_to_excel)
        btn_excel.setMinimumWidth(100)
        btn_excel.setStyleSheet(excel_button_style)
        layout.addWidget(btn_excel)
        
        # 새로고침 버튼
        btn_refresh = QPushButton("새로고침")
        btn_refresh.clicked.connect(self.load_products)
        btn_refresh.setMinimumWidth(80)
        btn_refresh.setStyleSheet(refresh_button_style)
        layout.addWidget(btn_refresh)
        
        return layout
    
    def create_table_area(self):
        """테이블 영역 생성"""
        layout = QVBoxLayout()
        
        label = QLabel("📋 제품 목록 (더블클릭으로 선택된 제품 수정)")
        font = QFont()
        font.setBold(True)
        font.setPointSize(11)
        label.setFont(font)
        label.setStyleSheet("color: #1976D2; padding: 5px;")
        layout.addWidget(label)
        
        # 테이블 생성
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(["ID", "제품명", "가격", "재고", "카테고리", "생성일"])
        self.table.setColumnWidth(0, 40)
        self.table.setColumnWidth(1, 150)
        self.table.setColumnWidth(2, 100)
        self.table.setColumnWidth(3, 80)
        self.table.setColumnWidth(4, 120)
        self.table.setColumnWidth(5, 150)
        self.table.setMinimumHeight(300)
        self.table.setSelectionBehavior(self.table.SelectRows)
        self.table.setSelectionMode(self.table.SingleSelection)
        
        # 테이블 스타일 설정
        self.table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #e0e0e0;
                border-radius: 5px;
                gridline-color: #e0e0e0;
                background-color: white;
            }
            QTableWidget::item {
                padding: 5px;
                border: none;
            }
            QTableWidget::item:selected {
                background-color: #E3F2FD;
                color: #1976D2;
            }
            QHeaderView::section {
                background-color: #1976D2;
                color: white;
                padding: 5px;
                border: none;
                font-weight: bold;
            }
        """)
        
        # 더블클릭 이벤트 연결
        self.table.doubleClicked.connect(self.load_row_to_input)
        
        layout.addWidget(self.table)
        
        return layout
    
    def load_products(self):
        """제품 목록 로드"""
        self.table.setRowCount(0)
        products = self.db.get_all_products()
        
        for row, product in enumerate(products):
            self.table.insertRow(row)
            
            # ID
            item = QTableWidgetItem(str(product['productID']))
            item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row, 0, item)
            
            # 제품명
            self.table.setItem(row, 1, QTableWidgetItem(product['productName']))
            
            # 가격
            item = QTableWidgetItem(f"{product['productPrice']:,}")
            item.setTextAlignment(Qt.AlignRight)
            self.table.setItem(row, 2, item)
            
            # 재고
            item = QTableWidgetItem(str(product['productStock']))
            item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row, 3, item)
            
            # 카테고리
            category = product['productCategory'] if product['productCategory'] else "-"
            self.table.setItem(row, 4, QTableWidgetItem(category))
            
            # 생성일
            self.table.setItem(row, 5, QTableWidgetItem(product['createdDate']))
    
    def load_row_to_input(self):
        """테이블의 더블클릭된 행을 입력 필드로 로드"""
        selected_row = self.table.currentRow()
        
        if selected_row < 0:
            return
        
        # 테이블에서 데이터 가져오기
        self.selected_product_id = int(self.table.item(selected_row, 0).text())
        product_name = self.table.item(selected_row, 1).text()
        product_price = int(self.table.item(selected_row, 2).text().replace(",", ""))
        product_stock = int(self.table.item(selected_row, 3).text())
        product_category = self.table.item(selected_row, 4).text()
        
        # 입력 필드에 데이터 설정
        self.input_name.setText(product_name)
        self.input_price.setValue(product_price)
        self.input_stock.setValue(product_stock)
        
        # 카테고리 설정
        index = self.input_category.findText(product_category)
        if index >= 0:
            self.input_category.setCurrentIndex(index)
        
        # 입력 필드에 포커스 설정
        self.input_name.setFocus()
    
    def add_product(self):
        """제품 추가"""
        name = self.input_name.text().strip()
        price = self.input_price.value()
        stock = self.input_stock.value()
        category = self.input_category.currentText()
        
        # 유효성 검사
        if not name:
            QMessageBox.warning(self, "입력 오류", "제품명을 입력해주세요.")
            return
        
        if price <= 0:
            QMessageBox.warning(self, "입력 오류", "가격을 입력해주세요.")
            return
        
        # 제품 추가
        try:
            self.db.insert_product(name, price, stock, category)
            self.load_products()
            
            # 입력 필드 초기화
            self.input_name.clear()
            self.input_price.setValue(0)
            self.input_stock.setValue(0)
            self.input_category.setCurrentIndex(0)
            self.selected_product_id = None
            
            QMessageBox.information(self, "성공", "제품이 추가되었습니다.")
        except Exception as e:
            QMessageBox.critical(self, "오류", f"제품 추가 실패: {str(e)}")
    
    def update_product(self):
        """제품 수정"""
        # 더블클릭으로 선택된 제품이 있는지 확인
        if self.selected_product_id is None:
            selected_row = self.table.currentRow()
            if selected_row < 0:
                QMessageBox.warning(self, "선택 오류", "수정할 제품을 선택해주세요.\n(테이블에서 더블클릭하거나 행을 선택하세요)")
                return
            product_id = int(self.table.item(selected_row, 0).text())
        else:
            product_id = self.selected_product_id
        
        name = self.input_name.text().strip()
        price = self.input_price.value()
        stock = self.input_stock.value()
        category = self.input_category.currentText()
        
        if not name:
            QMessageBox.warning(self, "입력 오류", "제품명을 입력해주세요.")
            return
        
        if price <= 0:
            QMessageBox.warning(self, "입력 오류", "가격을 입력해주세요.")
            return
        
        try:
            self.db.update_product(product_id, name, price, stock, category)
            self.load_products()
            
            self.input_name.clear()
            self.input_price.setValue(0)
            self.input_stock.setValue(0)
            self.input_category.setCurrentIndex(0)
            self.selected_product_id = None
            
            QMessageBox.information(self, "성공", "제품이 수정되었습니다.")
        except Exception as e:
            QMessageBox.critical(self, "오류", f"제품 수정 실패: {str(e)}")
    
    def delete_product(self):
        """제품 삭제"""
        # 더블클릭으로 선택된 제품이 있는지 확인
        if self.selected_product_id is None:
            selected_row = self.table.currentRow()
            if selected_row < 0:
                QMessageBox.warning(self, "선택 오류", "삭제할 제품을 선택해주세요.\n(테이블에서 더블클릭하거나 행을 선택하세요)")
                return
            product_id = int(self.table.item(selected_row, 0).text())
            product_name = self.table.item(selected_row, 1).text()
        else:
            product_id = self.selected_product_id
            product_name = self.input_name.text().strip()
        
        reply = QMessageBox.question(
            self, 
            "삭제 확인", 
            f"'{product_name}' 제품을 삭제하시겠습니까?",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            try:
                self.db.delete_product(product_id)
                self.load_products()
                
                self.input_name.clear()
                self.input_price.setValue(0)
                self.input_stock.setValue(0)
                self.input_category.setCurrentIndex(0)
                self.selected_product_id = None
                
                QMessageBox.information(self, "성공", "제품이 삭제되었습니다.")
            except Exception as e:
                QMessageBox.critical(self, "오류", f"제품 삭제 실패: {str(e)}")
    
    def save_to_excel(self):
        """엑셀 파일로 저장"""
        # 파일 저장 경로 선택
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "엑셀 파일 저장",
            f"products_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            products = self.db.get_all_products()
            
            if not products:
                QMessageBox.warning(self, "저장 실패", "저장할 제품이 없습니다.")
                return
            
            # 새로운 Workbook 생성
            wb = Workbook()
            ws = wb.active
            ws.title = "Products"
            
            # 헤더 설정
            headers = ["ID", "제품명", "가격", "재고", "카테고리", "생성일"]
            ws.append(headers)
            
            # 헤더 스타일 설정
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # 데이터 입력
            for product in products:
                ws.append([
                    product['productID'],
                    product['productName'],
                    product['productPrice'],
                    product['productStock'],
                    product['productCategory'] if product['productCategory'] else "-",
                    product['createdDate']
                ])
            
            # 열 너비 자동 조정
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 18
            ws.column_dimensions['F'].width = 20
            
            # 데이터 셀 스타일 설정
            center_alignment = Alignment(horizontal="center", vertical="center")
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
                for cell in row:
                    cell.alignment = center_alignment if cell.column in [1, 4, 6] else None
            
            # 가격 포맷 설정
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
                for cell in row:
                    cell.number_format = '#,##0'
            
            # 파일 저장
            wb.save(file_path)
            
            QMessageBox.information(
                self,
                "성공",
                f"엑셀 파일이 저장되었습니다.\n{file_path}"
            )
        
        except Exception as e:
            QMessageBox.critical(self, "오류", f"엑셀 저장 실패: {str(e)}")
    
    def closeEvent(self, event):
        """윈도우 종료 시 데이터베이스 연결 종료"""
        self.db.close()
        event.accept()


def main():
    app = QApplication(sys.argv)
    gui = ProductGUI()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
